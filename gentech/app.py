from __future__ import annotations

import json
import os
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List

from flask import Flask, abort, flash, redirect, render_template, request, send_file, url_for
from num2words import num2words
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
INVOICE_DIR = DATA_DIR / "invoices"
PDF_DIR = DATA_DIR / "generated" / "pdf"
XLSX_DIR = DATA_DIR / "generated" / "xlsx"

COMPANY_SETTINGS_PATH = DATA_DIR / "company_settings.json"
CUSTOMERS_PATH = DATA_DIR / "customers.json"
ITEMS_PATH = DATA_DIR / "items.json"
REFERENCE_OPTIONS_PATH = DATA_DIR / "references.json"
SEQUENCE_PATH = DATA_DIR / "sequence.json"
DEFAULT_LETTERHEAD_BASENAMES = (
    "letterhead.jpeg",
    "letterhead.jpg",
    "letterhead.png",
    "letterpad.jpeg",
    "letterpad.jpg",
    "letterpad.png",
)


app = Flask(__name__)
app.secret_key = "gentech-billing-secret"


def _load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        if default is None:
            return None
        _save_json(path, default)
        return default
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _save_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2)


def ensure_seed_data() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    INVOICE_DIR.mkdir(parents=True, exist_ok=True)
    PDF_DIR.mkdir(parents=True, exist_ok=True)
    XLSX_DIR.mkdir(parents=True, exist_ok=True)

    company = _load_json(
        COMPANY_SETTINGS_PATH,
        {
            "company_name": "GENTEC",
            "gstin": "33AXQPM7524G1Z8",
            "phone": "+91 8056789568",
            "header_phone_line": "Mob : 80567 89568, 91235 87990",
            "header_email": "E-mail : support@gentec.in",
            "header_website": "WebSite : www.gentecpowertechnologies.com",
            "invoice_prefix": "DGSP",
            "job_card_prefix": "GEN/CBE",
            "max_rows_per_invoice": 8,
            "max_line_items": 200,
            "delivery_terms_default": "work at site",
            "bank_details": [
                "A/c Name: Gentec",
                "Bank Name: Karur Vysya Bank",
                "A/c No: 1294135000006393",
                "Branch & IFS Code: Kovaipudur & KVBL0001294",
            ],
            "notes": [
                "Goods once sold will not be taken back.",
                "Interest @18% will be charged on unpaid amount after 45 days of invoice date.",
            ],
            "logo_path": "",
        },
    )
    # Use the attached letterhead file automatically if no custom logo path is configured.
    if not company.get("logo_path"):
        for basename in DEFAULT_LETTERHEAD_BASENAMES:
            default_letterhead = BASE_DIR / basename
            if default_letterhead.exists():
                company["logo_path"] = basename
                _save_json(COMPANY_SETTINGS_PATH, company)
                break
    company_defaults = {
        "header_phone_line": "Mob : 80567 89568, 91235 87990",
        "header_email": "E-mail : support@gentec.in",
        "header_website": "WebSite : www.gentecpowertechnologies.com",
    }
    updated = False
    for key, value in company_defaults.items():
        if not str(company.get(key, "")).strip():
            company[key] = value
            updated = True
    if updated:
        _save_json(COMPANY_SETTINGS_PATH, company)

    customers = _load_json(
        CUSTOMERS_PATH,
        [
            {
                "customer_id": "CUST001",
                "name": "Saaral Family Restaurant",
                "address": "34-F1, Karupparayan Kovil Street, Samalapuram, Palladam, Tiruppur - 641663",
                "gstin": "33FXEPS2789N1ZL",
                "reference_default": "Mahindra 15Kva",
            }
        ],
    )
    seeded_references = sorted(
        {
            str(customer.get("reference_default", "")).strip()
            for customer in customers
            if str(customer.get("reference_default", "")).strip()
        },
        key=str.lower,
    )
    _load_json(REFERENCE_OPTIONS_PATH, seeded_references)

    _load_json(
        ITEMS_PATH,
        [
            {"item_id": "ITEM001", "description": "Fuel filter", "hsn_sac": "8421", "default_unit_price": 220},
            {"item_id": "ITEM002", "description": "Lube Oil filter", "hsn_sac": "8421", "default_unit_price": 585},
            {"item_id": "ITEM003", "description": "Filter O ring", "hsn_sac": "8421", "default_unit_price": 125},
            {"item_id": "ITEM004", "description": "Banjo Washer", "hsn_sac": "7415", "default_unit_price": 22},
            {"item_id": "ITEM005", "description": "Stop Solenoid Coil", "hsn_sac": "8481", "default_unit_price": 4842},
        ],
    )

    _load_json(SEQUENCE_PATH, {"financial_year": "25-26", "current_number": 0})


ensure_seed_data()


def parse_money(value: str) -> float:
    if value is None:
        return 0.0
    clean = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return round(float(clean), 2)
    except ValueError:
        return 0.0


def parse_qty(value: str) -> float:
    if value is None:
        return 0.0
    try:
        return float(str(value).strip())
    except ValueError:
        return 0.0


def parse_bool(value: str | None) -> bool:
    return str(value).lower() in {"1", "true", "yes", "on"}


def normalize_reference_option(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def is_valid_reference_option(value: str) -> bool:
    normalized = normalize_reference_option(value)
    return normalized.lower() not in {"", "none", "manual", "manual entry", "manul entry"}


def get_reference_options(customers: List[Dict[str, Any]] | None = None) -> List[str]:
    customers = customers if customers is not None else get_customers()
    stored = _load_json(REFERENCE_OPTIONS_PATH, [])
    unique: Dict[str, str] = {}
    for raw in stored:
        normalized = normalize_reference_option(raw)
        if not is_valid_reference_option(normalized):
            continue
        unique.setdefault(normalized.lower(), normalized)
    for customer in customers:
        normalized = normalize_reference_option(customer.get("reference_default", ""))
        if not is_valid_reference_option(normalized):
            continue
        unique.setdefault(normalized.lower(), normalized)
    options = sorted(unique.values(), key=str.lower)
    save_reference_options(options)
    return options


def save_reference_options(values: List[str]) -> None:
    unique: Dict[str, str] = {}
    for raw in values:
        normalized = normalize_reference_option(raw)
        if not is_valid_reference_option(normalized):
            continue
        unique.setdefault(normalized.lower(), normalized)
    _save_json(REFERENCE_OPTIONS_PATH, sorted(unique.values(), key=str.lower))


def register_reference_option(value: Any) -> None:
    normalized = normalize_reference_option(value)
    if not is_valid_reference_option(normalized):
        return
    existing = _load_json(REFERENCE_OPTIONS_PATH, [])
    save_reference_options([*existing, normalized])


def resolve_letterhead_path(company: Dict[str, Any]) -> Path | None:
    configured = str(company.get("logo_path", "")).strip()
    candidates: List[Path] = []
    if configured:
        configured_path = Path(configured).expanduser()
        if not configured_path.is_absolute():
            configured_path = BASE_DIR / configured_path
        candidates.append(configured_path)
    candidates.extend(BASE_DIR / name for name in DEFAULT_LETTERHEAD_BASENAMES)
    for candidate in candidates:
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


def financial_year_for(invoice_date: date) -> str:
    if invoice_date.month >= 4:
        start = invoice_date.year
    else:
        start = invoice_date.year - 1
    end = start + 1
    return f"{start % 100:02d}-{end % 100:02d}"


def sequence_preview(invoice_date: date) -> tuple[int, str]:
    sequence = _load_json(SEQUENCE_PATH, {"financial_year": "25-26", "current_number": 0})
    fy = financial_year_for(invoice_date)
    if sequence.get("financial_year") != fy:
        return 1, fy
    return int(sequence.get("current_number", 0)) + 1, fy


def sequence_next(invoice_date: date) -> tuple[int, str]:
    sequence = _load_json(SEQUENCE_PATH, {"financial_year": "25-26", "current_number": 0})
    fy = financial_year_for(invoice_date)
    if sequence.get("financial_year") != fy:
        current = 1
    else:
        current = int(sequence.get("current_number", 0)) + 1
    _save_json(SEQUENCE_PATH, {"financial_year": fy, "current_number": current})
    return current, fy


def invoice_number(prefix: str, seq: int, fy: str) -> str:
    return f"{prefix}-{seq:02d}/{fy}"


def job_card_number(prefix: str, seq: int, fy: str) -> str:
    return f"{prefix}/{fy}/JC{seq:03d}"


def sanitize_amount_in_words(value: str) -> str:
    text = re.sub(r"^\s*(?:Rs\.?|₹)\s*", "", str(value or ""), flags=re.IGNORECASE)
    return text.strip().rstrip(".")


def amount_in_words(grand_total: float) -> str:
    rounded = int(round(grand_total))
    words = num2words(rounded, to="cardinal", lang="en_IN")
    return sanitize_amount_in_words(f"{words.title()} Only")


def resolve_pdf_font_profile() -> Dict[str, str]:
    cached = app.config.get("_pdf_font_profile")
    if isinstance(cached, dict):
        return cached

    profile = {
        "regular": "Times-Roman",
        "bold": "Times-Bold",
        "table_currency": "Rs",
        "words_currency": "Rs.",
    }
    font_pairs = [
        # Preferred: Bookman Old Style (project-local or OS fonts).
        (
            BASE_DIR / "fonts" / "Bookman Old Style.ttf",
            BASE_DIR / "fonts" / "Bookman Old Style Bold.ttf",
        ),
        (
            BASE_DIR / "fonts" / "BOOKOS.TTF",
            BASE_DIR / "fonts" / "BOOKOSB.TTF",
        ),
        (
            Path(r"C:\Windows\Fonts\BOOKOS.TTF"),
            Path(r"C:\Windows\Fonts\BOOKOSB.TTF"),
        ),
        (
            Path(r"C:\Windows\Fonts\bookos.ttf"),
            Path(r"C:\Windows\Fonts\bookosb.ttf"),
        ),
        (
            Path("/System/Library/Fonts/Supplemental/Bookman Old Style.ttf"),
            Path("/System/Library/Fonts/Supplemental/Bookman Old Style Bold.ttf"),
        ),
        (
            Path("/Library/Fonts/Bookman Old Style.ttf"),
            Path("/Library/Fonts/Bookman Old Style Bold.ttf"),
        ),
        # Fallbacks when Bookman is not present.
        (
            BASE_DIR / "fonts" / "Times New Roman.ttf",
            BASE_DIR / "fonts" / "Times New Roman Bold.ttf",
        ),
        (
            Path("/System/Library/Fonts/Supplemental/Times New Roman.ttf"),
            Path("/System/Library/Fonts/Supplemental/Times New Roman Bold.ttf"),
        ),
        (
            Path("/Library/Fonts/Times New Roman.ttf"),
            Path("/Library/Fonts/Times New Roman Bold.ttf"),
        ),
        (
            Path(r"C:\Windows\Fonts\times.ttf"),
            Path(r"C:\Windows\Fonts\timesbd.ttf"),
        ),
        (
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf"),
        ),
    ]
    for regular_path, bold_path in font_pairs:
        if not regular_path.exists() or not bold_path.exists():
            continue
        try:
            regular_name = "GentecPdfRegular"
            bold_name = "GentecPdfBold"
            registered = pdfmetrics.getRegisteredFontNames()
            if regular_name not in registered:
                pdfmetrics.registerFont(TTFont(regular_name, str(regular_path)))
            if bold_name not in registered:
                pdfmetrics.registerFont(TTFont(bold_name, str(bold_path)))
            profile = {
                "regular": regular_name,
                "bold": bold_name,
                "table_currency": "Rs",
                "words_currency": "Rs.",
            }
            break
        except Exception:
            continue

    app.config["_pdf_font_profile"] = profile
    return profile


def calculate_totals(rows: List[Dict[str, Any]], transport: float) -> Dict[str, float]:
    subtotal = round(sum(item["amount"] for item in rows), 2)
    cgst = round(subtotal * 0.09, 2)
    sgst = round(subtotal * 0.09, 2)
    grand_total = round(subtotal + cgst + sgst + transport, 2)
    return {
        "subtotal": subtotal,
        "cgst": cgst,
        "sgst": sgst,
        "transport": round(transport, 2),
        "grand_total": grand_total,
    }


def build_invoice_from_form(form: Any, item_map: Dict[str, Dict[str, Any]], company: Dict[str, Any]) -> Dict[str, Any]:
    customer_name = form.get("customer_name", "").strip()
    customer_address = form.get("customer_address", "").strip()
    customer_gstin = form.get("customer_gstin", "").strip()
    reference_direct = normalize_reference_option(form.get("customer_reference", ""))
    if reference_direct and reference_direct.lower() != "none":
        customer_reference = reference_direct
    else:
        # Backward compatibility for older forms still posting select/manual fields.
        reference_choice = form.get("customer_reference_select", "NONE").strip()
        reference_manual = normalize_reference_option(form.get("customer_reference_manual", ""))
        if reference_manual and reference_manual.lower() != "none":
            customer_reference = reference_manual
        elif reference_choice and reference_choice not in {"", "NONE", "MANUAL"}:
            customer_reference = reference_choice
        else:
            customer_reference = "None"

    invoice_date = datetime.strptime(form.get("invoice_date"), "%Y-%m-%d").date()
    job_card_date = datetime.strptime(form.get("job_card_date"), "%Y-%m-%d").date()
    order_date = datetime.strptime(form.get("order_date"), "%Y-%m-%d").date()

    descriptions = form.getlist("description")
    hsn_values = form.getlist("hsn_sac")
    qty_values = form.getlist("qty")
    unit_prices = form.getlist("unit_price")

    rows: List[Dict[str, Any]] = []
    for desc_raw, hsn_raw, qty_raw, price_raw in zip(descriptions, hsn_values, qty_values, unit_prices):
        description = desc_raw.strip()
        if not description:
            continue
        item = item_map.get(description.lower(), {})
        qty = parse_qty(qty_raw)
        hsn_sac = str(hsn_raw or "").strip() or str(item.get("hsn_sac", "")).strip()
        if str(price_raw).strip():
            unit_price = parse_money(price_raw)
        else:
            unit_price = parse_money(str(item.get("default_unit_price", "0")))
        amount = round(qty * unit_price, 2)
        rows.append(
            {
                "item_id": item.get("item_id", ""),
                "description": description,
                "hsn_sac": hsn_sac,
                "qty": qty,
                "unit_price": unit_price,
                "amount": amount,
            }
        )

    transport = parse_money(form.get("transport", "0"))
    totals = calculate_totals(rows, transport)

    if not rows:
        raise ValueError("At least one line item is required.")

    max_line_items = int(company.get("max_line_items", 200))
    if len(rows) > max_line_items:
        raise ValueError(f"Only {max_line_items} line items are allowed per invoice.")

    preview_seq, preview_fy = sequence_preview(invoice_date)

    return {
        "include_letterhead": parse_bool(form.get("include_letterhead")),
        "invoice_date": invoice_date.isoformat(),
        "job_card_date": job_card_date.isoformat(),
        "order_no": form.get("order_no", "Verbally").strip() or "Verbally",
        "order_date": order_date.isoformat(),
        "delivery_terms": form.get("delivery_terms", company.get("delivery_terms_default", "")).strip(),
        "customer": {
            "name": customer_name,
            "address": customer_address,
            "gstin": customer_gstin,
            "reference": customer_reference,
        },
        "rows": rows,
        "totals": totals,
        "amount_in_words": amount_in_words(totals["grand_total"]),
        "preview": {
            "sequence": preview_seq,
            "financial_year": preview_fy,
            "invoice_no": invoice_number(company.get("invoice_prefix", "DGSP"), preview_seq, preview_fy),
            "job_card_no": job_card_number(company.get("job_card_prefix", "GEN/CBE"), preview_seq, preview_fy),
        },
    }


def _wrap_canvas_text(c: canvas.Canvas, text: str, max_width: float, font_name: str, font_size: float) -> List[str]:
    words = str(text or "").split()
    if not words:
        return [""]
    lines: List[str] = []
    current = ""
    for word in words:
        candidate = word if not current else f"{current} {word}"
        if c.stringWidth(candidate, font_name, font_size) <= max_width:
            current = candidate
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


def _draw_wrapped_block(
    c: canvas.Canvas,
    text: str,
    x: float,
    start_y: float,
    max_width: float,
    min_y: float,
    font_name: str,
    font_size: float,
    line_height: float,
) -> float:
    y = start_y
    c.setFont(font_name, font_size)
    for line in _wrap_canvas_text(c, text, max_width, font_name, font_size):
        if y < min_y:
            break
        c.drawString(x, y, line)
        y -= line_height
    return y


def generate_pdf(invoice: Dict[str, Any], company: Dict[str, Any], output_path: Path) -> None:
    c = canvas.Canvas(str(output_path), pagesize=A4)
    width, height = A4
    font_profile = resolve_pdf_font_profile()
    pdf_font_regular = font_profile["regular"]
    pdf_font_bold = font_profile["bold"]
    table_currency = font_profile["table_currency"]
    words_currency = font_profile["words_currency"]
    words_value = sanitize_amount_in_words(invoice.get("amount_in_words", ""))
    margin = 22
    top_y = height - margin
    letterhead_path = resolve_letterhead_path(company)
    # Keep invoice table fully below the pre-printed logo/header area.
    stationery_top_offset = 94
    first_page_capacity = int(company.get("max_rows_per_invoice", 8))
    continued_page_capacity = 22

    def draw_background() -> None:
        if not letterhead_path:
            return
        c.drawImage(
            str(letterhead_path),
            0,
            0,
            width=width,
            height=height,
            preserveAspectRatio=False,
            mask="auto",
        )
        if not invoice.get("include_letterhead"):
            mask_height = stationery_top_offset + 22
            c.setFillColor(colors.white)
            c.rect(0, height - mask_height, width, mask_height, stroke=0, fill=1)
            c.setFillColor(colors.black)

    draw_background()

    page_rows = invoice["rows"][:first_page_capacity]
    continued_rows = invoice["rows"][first_page_capacity:]
    item_rows = len(page_rows)

    h_title = 21
    h_gst = 25
    h_customer = 140
    h_table_header = 27
    h_item = 22
    h_total = 19
    base_outer_bottom = margin + 72
    lift_per_missing_row = 8
    missing_rows = max(0, first_page_capacity - item_rows)

    outer_left = margin + 20
    outer_right = width - margin - 20
    outer_top = top_y - (stationery_top_offset if letterhead_path else 54)
    layout_bottom = base_outer_bottom + (missing_rows * lift_per_missing_row)
    outer_width = outer_right - outer_left
    split_x = round(outer_left + (outer_width * 0.50))

    c.setLineWidth(0.85)

    customer = invoice["customer"]

    y = outer_top

    # Title row
    y_title_bottom = y - h_title
    c.setFillColor(colors.HexColor("#ececec"))
    c.rect(outer_left, y_title_bottom, outer_width, h_title, stroke=0, fill=1)
    c.setFillColor(colors.black)
    c.line(outer_left, y_title_bottom, outer_right, y_title_bottom)
    c.setFont(pdf_font_bold, 12.5)
    c.drawCentredString((outer_left + outer_right) / 2, y - 14, "TAX INVOICE")
    y = y_title_bottom

    # GST / recipient row
    y_gst_bottom = y - h_gst
    c.line(outer_left, y_gst_bottom, outer_right, y_gst_bottom)
    c.line(split_x, y, split_x, y_gst_bottom)
    c.setFont(pdf_font_bold, 10.2)
    c.drawString(outer_left + 4, y - 16, f"GSTIN: {company.get('gstin', '')}")
    c.drawCentredString((split_x + outer_right) / 2, y - 16, "Original for Recipient")
    y = y_gst_bottom

    if letterhead_path and invoice.get("include_letterhead"):
        contact_x = outer_left - 2
        contact_top = outer_top + 64
        contact_width = 270
        contact_height = 56
        c.setFillColor(colors.white)
        c.rect(contact_x - 2, contact_top - contact_height, contact_width, contact_height, stroke=0, fill=1)
        c.setFillColor(colors.black)
        c.setFont(pdf_font_bold, 9.2)
        c.drawString(contact_x, contact_top - 12, company.get("header_phone_line", ""))
        c.drawString(contact_x, contact_top - 25, company.get("header_email", ""))
        c.drawString(contact_x, contact_top - 38, company.get("header_website", ""))

    # Customer and right metadata block
    y_customer_bottom = y - h_customer
    c.line(outer_left, y_customer_bottom, outer_right, y_customer_bottom)
    c.line(split_x, y, split_x, y_customer_bottom)

    meta_rows = [
        ("Invoice No", invoice["invoice_no"]),
        ("Invoice Date", invoice["invoice_date"]),
        ("Job Card No", invoice["job_card_no"]),
        ("Job Card Date", invoice["job_card_date"]),
        ("Cust's Order No", invoice["order_no"]),
        ("Date", invoice["order_date"]),
        ("Delivery Terms", invoice["delivery_terms"]),
    ]
    right_row_h = h_customer / len(meta_rows)
    c.setFont(pdf_font_bold, 11)
    c.drawString(outer_left + 4, y - 18, "Customer Details")

    c.setFont(pdf_font_bold, 11)
    c.drawString(outer_left + 5, y - 40, customer["name"])

    address_tokens = [segment.strip() for segment in str(customer["address"]).split(",") if segment.strip()]
    if len(address_tokens) >= 3:
        address_lines = [address_tokens[0], address_tokens[1], ", ".join(address_tokens[2:])]
    elif len(address_tokens) == 2:
        address_lines = [address_tokens[0], address_tokens[1], ""]
    else:
        address_lines = _wrap_canvas_text(c, customer["address"], split_x - outer_left - 16, pdf_font_regular, 10.8)
        while len(address_lines) < 3:
            address_lines.append("")
        address_lines = address_lines[:3]

    c.setFont(pdf_font_regular, 10.8)
    address_y = y - 58
    for line in address_lines:
        if line:
            c.drawString(outer_left + 5, address_y, line)
        address_y -= 13

    c.drawString(outer_left + 5, y_customer_bottom + 28, f"GSTIN: {customer['gstin']}")
    c.drawString(outer_left + 5, y_customer_bottom + 12, f"Ref: {customer['reference']}")

    current_top = y
    value_split_x = split_x + 136
    for index, (label, value) in enumerate(meta_rows):
        current_bottom = y - ((index + 1) * right_row_h)
        c.line(split_x, current_bottom, outer_right, current_bottom)
        c.line(value_split_x, current_top, value_split_x, current_bottom)
        c.setFont(pdf_font_bold, 10.6)
        c.drawRightString(value_split_x - 6, current_top - 15, f"{label}:")
        c.setFont(pdf_font_regular, 10.2)
        c.drawString(value_split_x + 6, current_top - 15, str(value))
        current_top = current_bottom

    y = y_customer_bottom

    # Item table
    y_header_bottom = y - h_table_header
    c.line(outer_left, y_header_bottom, outer_right, y_header_bottom)

    col_sl = outer_left + 34
    col_desc = col_sl + 188
    col_qty = split_x
    col_hsn = col_qty + 58
    col_unit = value_split_x
    for xpos in [col_sl, col_desc, col_qty, col_hsn, col_unit]:
        c.line(xpos, y, xpos, y_header_bottom - (item_rows * h_item))

    c.setFont(pdf_font_bold, 10.8)
    c.drawCentredString((outer_left + col_sl) / 2, y - 17, "Sl.No")
    c.drawCentredString((col_sl + col_desc) / 2, y - 17, "Description")
    c.drawCentredString((col_desc + col_qty) / 2, y - 17, "Qty")
    c.drawCentredString((col_qty + col_hsn) / 2, y - 17, "HSN/SAC")
    c.drawCentredString((col_hsn + col_unit) / 2, y - 15, "Unit Price")
    c.drawCentredString((col_hsn + col_unit) / 2, y - 24, f"in {table_currency}")
    c.drawCentredString((col_unit + outer_right) / 2, y - 15, "Amount")
    c.drawCentredString((col_unit + outer_right) / 2, y - 24, f"in {table_currency}")

    item_y = y_header_bottom
    c.setFont(pdf_font_regular, 11)
    for index, row in enumerate(page_rows, start=1):
        next_y = item_y - h_item
        c.line(outer_left, next_y, outer_right, next_y)
        c.drawCentredString((outer_left + col_sl) / 2, item_y - 15, str(index))
        c.drawCentredString((col_sl + col_desc) / 2, item_y - 16, row["description"])
        c.drawCentredString((col_desc + col_qty) / 2, item_y - 15, f"{row['qty']:g}")
        c.drawCentredString((col_qty + col_hsn) / 2, item_y - 16, row["hsn_sac"])
        c.drawRightString(col_unit - 7, item_y - 16, f"{row['unit_price']:.2f}")
        c.drawRightString(outer_right - 5, item_y - 16, f"{row['amount']:.2f}")
        item_y = next_y

    totals = invoice["totals"]
    totals_rows = [
        ("Total Amount", totals["subtotal"], True),
        ("CGST 9%", totals["cgst"], True),
        ("SGST 9%", totals["sgst"], True),
        ("Transport", totals["transport"], False),
        ("Grand Total", totals["grand_total"], True),
    ]
    totals_top = item_y
    c.line(col_unit, totals_top, col_unit, totals_top - (h_total * len(totals_rows)))
    total_y = totals_top
    for label, value, bold in totals_rows:
        next_y = total_y - h_total
        c.line(outer_left, next_y, outer_right, next_y)
        c.setFont(pdf_font_bold if bold else pdf_font_regular, 11.2 if bold else 10.8)
        c.drawRightString(col_unit - 5, total_y - 14, label)
        c.drawRightString(outer_right - 5, total_y - 14, f"{value:.2f}")
        total_y = next_y

    # Footer details block
    footer_top = total_y
    footer_min_y = layout_bottom + 12
    left_block_right = split_x + 40
    amount_words_y = footer_top - 16
    amount_label = "Amount In words:"
    c.setFont(pdf_font_regular, 11)
    c.drawString(outer_left + 4, amount_words_y, amount_label)
    amount_label_width = c.stringWidth(f"{amount_label} ", pdf_font_regular, 11)
    c.setFont(pdf_font_bold, 11)
    c.drawString(outer_left + 4 + amount_label_width, amount_words_y, f"{words_currency} {words_value}.")
    c.setFont(pdf_font_bold, 11.8)
    c.drawString(outer_left + 4, footer_top - 34, "Mode of Payment:Cash/Cheque to GENTEC.")

    note_heading_y = footer_top - 52
    notes_y = note_heading_y
    c.setFont(pdf_font_bold, 11)
    c.drawString(outer_left + 4, notes_y, "Note:")
    notes_y -= 16
    note_prefixes = ["i", "ii", "iii", "iv", "v", "vi"]
    for idx, note in enumerate(company.get("notes", []), start=1):
        prefix = note_prefixes[idx - 1] if idx <= len(note_prefixes) else str(idx)
        notes_y = _draw_wrapped_block(
            c,
            f"{prefix}) {note}",
            outer_left + 4,
            notes_y,
            left_block_right - outer_left - 8,
            footer_min_y + 48,
            pdf_font_regular,
            10.6,
            12,
        ) - 2

    c.setFont(pdf_font_bold, 11)
    bank_heading_y = max(footer_min_y + 28, notes_y)
    c.drawString(outer_left + 4, bank_heading_y, "Bank Details:-")
    bank_line_y = max(footer_min_y + 20, bank_heading_y - 14)
    last_bank_line_y = bank_line_y
    for line in company.get("bank_details", []):
        c.setFont(pdf_font_bold, 10.8)
        c.drawString(outer_left + 4, bank_line_y, line)
        last_bank_line_y = bank_line_y
        bank_line_y -= 11

    # Right-side signature block.
    c.setFont(pdf_font_bold, 13)
    c.drawRightString(outer_right - 4, note_heading_y, "For GENTEC")
    mobile_y = max(last_bank_line_y, layout_bottom + 18)
    c.setFont(pdf_font_bold, 12.2)
    c.drawRightString(outer_right - 4, mobile_y + 20, "Authorised Signatory")
    c.setFont(pdf_font_regular, 12)
    c.drawRightString(outer_right - 4, mobile_y, f"Mobile: {company.get('phone', '')}")

    border_bottom = max(layout_bottom, mobile_y - 10)
    c.setLineWidth(1.1)
    c.rect(outer_left, border_bottom, outer_width, outer_top - border_bottom)
    c.setLineWidth(0.85)

    c.showPage()

    # Continue only the item table on additional pages.
    if continued_rows:
        col_sl_w = 36
        col_desc_w = 165
        col_qty_w = 42
        col_hsn_w = 64
        col_unit_w = 98
        current_index = first_page_capacity + 1
        for chunk_start in range(0, len(continued_rows), continued_page_capacity):
            chunk = continued_rows[chunk_start : chunk_start + continued_page_capacity]
            draw_background()
            page_outer_left = outer_left
            page_outer_right = outer_right
            page_top = top_y - (stationery_top_offset if letterhead_path else 54)
            c.setFont(pdf_font_bold, 12)
            c.drawString(page_outer_left, page_top, "Page continued...")

            table_top = page_top - 14
            table_header_h = 27
            row_h = 22
            table_height = table_header_h + (row_h * len(chunk))
            table_bottom = table_top - table_height
            c.rect(page_outer_left, table_bottom, page_outer_right - page_outer_left, table_height)

            c.line(page_outer_left, table_top - table_header_h, page_outer_right, table_top - table_header_h)
            c1 = page_outer_left + 34
            c2 = c1 + 188
            c3 = split_x
            c4 = c3 + 58
            c5 = value_split_x
            for xpos in [c1, c2, c3, c4, c5]:
                c.line(xpos, table_top, xpos, table_bottom)

            c.setFont(pdf_font_bold, 10.8)
            c.drawCentredString((page_outer_left + c1) / 2, table_top - 17, "Sl.No")
            c.drawCentredString((c1 + c2) / 2, table_top - 17, "Description")
            c.drawCentredString((c2 + c3) / 2, table_top - 17, "Qty")
            c.drawCentredString((c3 + c4) / 2, table_top - 17, "HSN/SAC")
            c.drawCentredString((c4 + c5) / 2, table_top - 15, "Unit Price")
            c.drawCentredString((c4 + c5) / 2, table_top - 24, f"in {table_currency}")
            c.drawCentredString((c5 + page_outer_right) / 2, table_top - 15, "Amount")
            c.drawCentredString((c5 + page_outer_right) / 2, table_top - 24, f"in {table_currency}")

            c.setFont(pdf_font_regular, 11)
            row_y = table_top - table_header_h
            for row in chunk:
                next_row_y = row_y - row_h
                c.line(page_outer_left, next_row_y, page_outer_right, next_row_y)
                c.drawCentredString((page_outer_left + c1) / 2, row_y - 15, str(current_index))
                c.drawCentredString((c1 + c2) / 2, row_y - 16, row["description"])
                c.drawCentredString((c2 + c3) / 2, row_y - 15, f"{row['qty']:g}")
                c.drawCentredString((c3 + c4) / 2, row_y - 16, row["hsn_sac"])
                c.drawRightString(c5 - 7, row_y - 16, f"{row['unit_price']:.2f}")
                c.drawRightString(page_outer_right - 5, row_y - 16, f"{row['amount']:.2f}")
                row_y = next_row_y
                current_index += 1

            c.showPage()

    c.save()


def generate_excel(invoice: Dict[str, Any], company: Dict[str, Any], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    thin = Side(border_style="thin", color="000000")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    letterhead_path = resolve_letterhead_path(company)

    # Fixed-height header band reserved in both modes.
    for row in range(1, 5):
        ws.row_dimensions[row].height = 20
    ws.merge_cells("A1:F4")
    if invoice.get("include_letterhead"):
        if letterhead_path:
            img = XLImage(str(letterhead_path))
            img.width = 520
            img.height = 78
            ws.add_image(img, "A1")
        else:
            ws["A1"] = f"{company.get('company_name', 'GENTEC')}\nGSTIN: {company.get('gstin', '')}\nMobile: {company.get('phone', '')}"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A1"].font = Font(size=13, bold=True)
    ws["A1"].border = cell_border

    ws.merge_cells("A5:F5")
    ws["A5"] = "TAX INVOICE"
    ws["A5"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A5"].font = Font(size=14, bold=True)
    ws["A5"].border = cell_border

    ws.merge_cells("A6:C6")
    ws.merge_cells("D6:F6")
    ws["A6"] = "Customer Details"
    ws["D6"] = "Original for Recipient"
    ws["A6"].font = ws["D6"].font = Font(bold=True)

    customer = invoice["customer"]
    ws.merge_cells("A7:C7")
    ws["A7"] = customer["name"]
    ws.merge_cells("D7:F7")
    ws["D7"] = f"Invoice No: {invoice['invoice_no']}"

    ws.merge_cells("A8:C8")
    ws["A8"] = customer["address"]
    ws.merge_cells("D8:F8")
    ws["D8"] = f"Invoice Date: {invoice['invoice_date']}"

    ws.merge_cells("A9:C9")
    ws["A9"] = f"GSTIN: {customer['gstin']} | Ref: {customer['reference']}"
    ws.merge_cells("D9:F9")
    ws["D9"] = f"Job Card No: {invoice['job_card_no']}"

    ws.merge_cells("D10:F10")
    ws["D10"] = f"Job Card Date: {invoice['job_card_date']}"

    ws.merge_cells("D11:F11")
    ws["D11"] = f"Cust Order No: {invoice['order_no']}  Date: {invoice['order_date']}"

    table_start = 13
    headers = ["Sl.No", "Description", "Qty", "HSN/SAC", "Unit Price (Rs)", "Amount (Rs)"]
    for index, label in enumerate(headers, start=1):
        cell = ws.cell(table_start, index, label)
        cell.font = Font(bold=True)
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_no = table_start + 1
    for idx, row in enumerate(invoice["rows"], start=1):
        ws.cell(row_no, 1, idx)
        ws.cell(row_no, 2, row["description"])
        ws.cell(row_no, 3, row["qty"])
        ws.cell(row_no, 4, row["hsn_sac"])
        ws.cell(row_no, 5, row["unit_price"])
        ws.cell(row_no, 6, row["amount"])
        for col in range(1, 7):
            ws.cell(row_no, col).border = cell_border
        row_no += 1

    max_rows = int(company.get("max_rows_per_invoice", 8))
    for _ in range(len(invoice["rows"]), max_rows):
        for col in range(1, 7):
            ws.cell(row_no, col, "").border = cell_border
        row_no += 1

    totals = invoice["totals"]
    summary = [
        ("Total Amount", totals["subtotal"]),
        ("CGST 9%", totals["cgst"]),
        ("SGST 9%", totals["sgst"]),
        ("Transport", totals["transport"]),
        ("Grand Total", totals["grand_total"]),
    ]
    for label, value in summary:
        ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=5)
        ws.cell(row_no, 1, label).alignment = Alignment(horizontal="right")
        is_bold = label in {"CGST 9%", "SGST 9%", "Grand Total"}
        ws.cell(row_no, 1).font = Font(bold=is_bold)
        ws.cell(row_no, 6, value).font = Font(bold=is_bold)
        for col in range(1, 7):
            ws.cell(row_no, col).border = cell_border
        row_no += 1

    ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=6)
    words_value = sanitize_amount_in_words(invoice.get("amount_in_words", ""))
    ws.cell(row_no, 1, f"Amount in words: Rs. {words_value}")

    wb.save(output_path)


def get_customers() -> List[Dict[str, Any]]:
    return _load_json(CUSTOMERS_PATH, [])


def get_items() -> List[Dict[str, Any]]:
    return _load_json(ITEMS_PATH, [])


def get_company() -> Dict[str, Any]:
    return _load_json(COMPANY_SETTINGS_PATH, {})


def normalize_invoice_for_template(invoice: Dict[str, Any]) -> Dict[str, Any]:
    invoice["include_letterhead"] = bool(invoice.get("include_letterhead"))
    invoice["amount_in_words"] = sanitize_amount_in_words(invoice.get("amount_in_words", ""))
    return invoice


@app.route("/assets/letterhead")
def letterhead_asset() -> Any:
    path = resolve_letterhead_path(get_company())
    if not path:
        abort(404)
    return send_file(path)


@app.route("/")
def home() -> Any:
    return redirect(url_for("new_invoice"))


@app.route("/invoice/new")
def new_invoice() -> Any:
    customers = get_customers()
    items = get_items()
    company = get_company()
    reference_options = get_reference_options(customers)
    today = date.today().isoformat()
    return render_template(
        "invoice_form.html",
        customers=customers,
        items=items,
        reference_options=reference_options,
        company=company,
        max_line_items=int(company.get("max_line_items", 200)),
        today=today,
        selected_customer=request.args.get("customer_id", ""),
    )


@app.route("/invoice/preview", methods=["POST"])
def preview_invoice() -> Any:
    try:
        company = get_company()
        if request.form.get("payload_json"):
            invoice = json.loads(request.form.get("payload_json", "{}"))
            invoice["include_letterhead"] = parse_bool(request.form.get("include_letterhead"))
        else:
            item_map = {str(item.get("description", "")).strip().lower(): item for item in get_items()}
            invoice = build_invoice_from_form(request.form, item_map, company)
            register_reference_option(invoice.get("customer", {}).get("reference", ""))

        invoice = normalize_invoice_for_template(invoice)
        payload_json = json.dumps(invoice)
        return render_template(
            "invoice_preview.html",
            invoice=invoice,
            company=company,
            payload_json=payload_json,
            has_letterhead=resolve_letterhead_path(company) is not None,
        )
    except ValueError as exc:
        flash(str(exc), "error")
        return redirect(url_for("new_invoice"))


@app.route("/invoice/save", methods=["POST"])
def save_invoice() -> Any:
    company = get_company()
    payload_json = request.form.get("payload_json", "{}")
    invoice = json.loads(payload_json)
    invoice["include_letterhead"] = parse_bool(request.form.get("include_letterhead"))
    invoice = normalize_invoice_for_template(invoice)
    register_reference_option(invoice.get("customer", {}).get("reference", ""))

    invoice_date = datetime.strptime(invoice["invoice_date"], "%Y-%m-%d").date()
    seq, fy = sequence_next(invoice_date)

    inv_no = invoice_number(company.get("invoice_prefix", "DGSP"), seq, fy)
    jc_no = job_card_number(company.get("job_card_prefix", "GEN/CBE"), seq, fy)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    invoice_id = f"{timestamp}-{seq:03d}"
    safe_name = inv_no.replace("/", "-")

    invoice["invoice_id"] = invoice_id
    invoice["financial_year"] = fy
    invoice["sequence"] = seq
    invoice["invoice_no"] = inv_no
    invoice["job_card_no"] = jc_no
    invoice["created_at"] = datetime.now().isoformat(timespec="seconds")

    pdf_path = PDF_DIR / f"{safe_name}.pdf"
    xlsx_path = XLSX_DIR / f"{safe_name}.xlsx"

    generate_pdf(invoice, company, pdf_path)
    generate_excel(invoice, company, xlsx_path)

    invoice["pdf_file"] = str(pdf_path.relative_to(BASE_DIR))
    invoice["xlsx_file"] = str(xlsx_path.relative_to(BASE_DIR))

    _save_json(INVOICE_DIR / f"{invoice_id}.json", invoice)
    flash(f"Invoice {inv_no} saved successfully.", "success")
    return redirect(url_for("history"))


@app.route("/customers/new", methods=["GET", "POST"])
def add_customer() -> Any:
    if request.method == "POST":
        customers = get_customers()
        next_id = f"CUST{len(customers) + 1:03d}"
        customer = {
            "customer_id": next_id,
            "name": request.form.get("name", "").strip(),
            "address": request.form.get("address", "").strip(),
            "gstin": request.form.get("gstin", "").strip(),
            "reference_default": request.form.get("reference_default", "").strip(),
        }
        customers.append(customer)
        _save_json(CUSTOMERS_PATH, customers)
        register_reference_option(customer.get("reference_default", ""))
        flash("Customer added successfully.", "success")
        return redirect(url_for("new_invoice", customer_id=next_id))
    return render_template("customer_new.html")


@app.route("/items/new", methods=["GET", "POST"])
def add_item() -> Any:
    if request.method == "POST":
        items = get_items()
        next_id = f"ITEM{len(items) + 1:03d}"
        item = {
            "item_id": next_id,
            "description": request.form.get("description", "").strip(),
            "hsn_sac": request.form.get("hsn_sac", "").strip(),
            "default_unit_price": parse_money(request.form.get("default_unit_price", "0")),
        }
        if not item["description"]:
            flash("Item description is required.", "error")
            return render_template("item_new.html")
        items.append(item)
        _save_json(ITEMS_PATH, items)
        flash("Item added successfully.", "success")
        return redirect(url_for("new_invoice"))
    return render_template("item_new.html")


@app.route("/history")
def history() -> Any:
    invoices: List[Dict[str, Any]] = []
    for path in sorted(INVOICE_DIR.glob("*.json"), reverse=True):
        data = _load_json(path, {})
        invoices.append(data)
    return render_template("history.html", invoices=invoices)


@app.route("/invoice/<invoice_id>/pdf")
def download_pdf(invoice_id: str) -> Any:
    payload = _load_json(INVOICE_DIR / f"{invoice_id}.json", None)
    if not payload:
        flash("Invoice not found.", "error")
        return redirect(url_for("history"))
    return send_file(BASE_DIR / payload["pdf_file"], as_attachment=True)


@app.route("/invoice/<invoice_id>/xlsx")
def download_xlsx(invoice_id: str) -> Any:
    payload = _load_json(INVOICE_DIR / f"{invoice_id}.json", None)
    if not payload:
        flash("Invoice not found.", "error")
        return redirect(url_for("history"))
    return send_file(BASE_DIR / payload["xlsx_file"], as_attachment=True)


@app.route("/invoice/<invoice_id>/delete", methods=["POST"])
def delete_invoice(invoice_id: str) -> Any:
    invoice_path = INVOICE_DIR / f"{invoice_id}.json"
    payload = _load_json(invoice_path, None)
    if not payload:
        flash("Invoice not found.", "error")
        return redirect(url_for("history"))

    base_resolved = BASE_DIR.resolve()
    for key in ("pdf_file", "xlsx_file"):
        rel_path = payload.get(key)
        if not rel_path:
            continue
        file_path = (BASE_DIR / rel_path).resolve()
        if base_resolved in file_path.parents and file_path.exists():
            file_path.unlink()

    if invoice_path.exists():
        invoice_path.unlink()

    flash(f"Invoice {payload.get('invoice_no', invoice_id)} deleted.", "success")
    return redirect(url_for("history"))


if __name__ == "__main__":
    ensure_seed_data()
    port = int(os.getenv("PORT", "5000"))
    debug = os.getenv("FLASK_DEBUG", "1") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
